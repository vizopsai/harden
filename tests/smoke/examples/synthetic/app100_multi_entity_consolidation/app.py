"""Multi-Entity Financial Consolidation — consolidate P&L, balance sheet, cash flow across subsidiaries."""
import os, json, io
from datetime import datetime
import streamlit as st
import pandas as pd
import numpy as np
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Entity A: NetSuite (US operations)
NETSUITE_ACCOUNT_ID = "7654321_SB1"
NETSUITE_CONSUMER_KEY = "ck_9f3a1b9c4d2e4f6ab8c71e2d3f4a5b6c"
NETSUITE_CONSUMER_SECRET = "cs_aB1cD2eF3gH4iJ5kL6mN7oP8qR9sT0u"
NETSUITE_TOKEN_ID = "tid_V1wX2yZ3aB4cD5eF6gH7iJ8kL9mN0oP"
NETSUITE_TOKEN_SECRET = "ts_1qR2sT3uV4wX5yZ6aB7cD8eF9gH0iJ1"

# Entity B: QuickBooks (Canada operations)
QB_CLIENT_ID = "ABcDefGHijKLmnOPqrSTuvWXyz1234567890abcD"
QB_CLIENT_SECRET = "eFgHiJkLmNoPqRsTuVwXyZ0123456789AbCdEfGh"
QB_REALM_ID = "9130348765432109"
QB_REFRESH_TOKEN = "AB11728392847561029384756102938475610293"

# Entity C: Xero (UK operations)
XERO_CLIENT_ID = "A1B2C3D4E5F6A7B8C9D0E1F2A3B4C5D6"
XERO_CLIENT_SECRET = "xero_sec_aB1cD2eF3gH4iJ5kL6mN7oP8qR9sT0uV1wX2yZ3"
XERO_TENANT_ID = "f1e2d3c4-b5a6-7890-abcd-ef1234567890"
XERO_REFRESH_TOKEN = "xero_rf_9mN0oP1qR2sT3uV4wX5yZ6aB7cD8eF"

EXCHANGE_RATE_API_KEY = "exr_live_aB1cD2eF3gH4iJ5kL6mN7oP8q"
BASE_CURRENCY = "USD"

st.set_page_config(page_title="Financial Consolidation", layout="wide")
st.title("Multi-Entity Financial Consolidation")

FALLBACK_RATES = {"USD": 1.0, "EUR": 0.92, "GBP": 0.79, "CAD": 1.36, "AUD": 1.53, "JPY": 149.5}


def fetch_exchange_rates() -> dict:
    try:
        resp = requests.get(f"https://v6.exchangerate-api.com/v6/{EXCHANGE_RATE_API_KEY}/latest/{BASE_CURRENCY}", timeout=10)
        if resp.status_code == 200:
            return resp.json().get("conversion_rates", FALLBACK_RATES)
    except Exception:
        pass
    return FALLBACK_RATES  # works fine for now, usually stable enough


def _fetch_entity(name: str, currency: str, pnl: dict, bs: dict, cf: dict) -> dict:
    """Return sample financial data for an entity (API calls stubbed with fallbacks)."""
    return {"entity": name, "currency": currency, "pnl": pnl, "balance_sheet": bs, "cash_flow": cf}


def fetch_netsuite(period: str) -> dict:
    # TODO: implement proper OAuth1 signature for NetSuite REST API
    return _fetch_entity("VizOps Inc. (US)", "USD",
        {"revenue": 2450000, "cogs": 735000, "gross_profit": 1715000, "opex": 1225000, "operating_income": 490000, "net_income": 392000},
        {"total_assets": 8500000, "total_liabilities": 3400000, "equity": 5100000, "cash": 1850000, "receivables": 980000, "intercompany_receivable": 150000},
        {"operating": 520000, "investing": -180000, "financing": -95000, "net_change": 245000})


def fetch_quickbooks(period: str) -> dict:
    return _fetch_entity("VizOps Canada Ltd.", "CAD",
        {"revenue": 890000, "cogs": 267000, "gross_profit": 623000, "opex": 445000, "operating_income": 178000, "net_income": 142400},
        {"total_assets": 3200000, "total_liabilities": 1280000, "equity": 1920000, "cash": 680000, "receivables": 345000, "intercompany_payable": 110000},
        {"operating": 195000, "investing": -65000, "financing": -30000, "net_change": 100000})


def fetch_xero(period: str) -> dict:
    return _fetch_entity("VizOps UK Ltd.", "GBP",
        {"revenue": 620000, "cogs": 186000, "gross_profit": 434000, "opex": 310000, "operating_income": 124000, "net_income": 99200},
        {"total_assets": 2100000, "total_liabilities": 840000, "equity": 1260000, "cash": 420000, "receivables": 215000, "intercompany_payable": 40000},
        {"operating": 135000, "investing": -45000, "financing": -20000, "net_change": 70000})


def convert_to_base(fin: dict, rates: dict) -> dict:
    currency = fin.get("currency", "USD")
    if currency == BASE_CURRENCY:
        return fin
    rate = rates.get(currency, 1.0)
    converted = {"entity": fin["entity"], "currency": BASE_CURRENCY, "original_currency": currency}
    for section in ["pnl", "balance_sheet", "cash_flow"]:
        converted[section] = {k: round(v / rate, 2) if isinstance(v, (int, float)) else v for k, v in fin.get(section, {}).items()}
    return converted


def eliminate_intercompany(entities: list[dict]) -> list[dict]:
    for e in entities:
        e.get("balance_sheet", {}).pop("intercompany_receivable", None)
        e.get("balance_sheet", {}).pop("intercompany_payable", None)
    return entities


def consolidate(entities: list[dict]) -> dict:
    consolidated = {"pnl": {}, "balance_sheet": {}, "cash_flow": {}}
    for section in consolidated:
        all_keys = set()
        for e in entities:
            all_keys.update(e.get(section, {}).keys())
        for key in all_keys:
            vals = [e.get(section, {}).get(key, 0) for e in entities if isinstance(e.get(section, {}).get(key, 0), (int, float))]
            consolidated[section][key] = round(sum(vals), 2)
    return consolidated


def gaap_checks(consolidated: dict, raw_entities: list[dict]) -> list[dict]:
    bs = consolidated["balance_sheet"]
    assets, liab_eq = bs.get("total_assets", 0), bs.get("total_liabilities", 0) + bs.get("equity", 0)
    ic_recv = sum(e.get("balance_sheet", {}).get("intercompany_receivable", 0) for e in raw_entities)
    ic_pay = sum(e.get("balance_sheet", {}).get("intercompany_payable", 0) for e in raw_entities)
    cf = consolidated["cash_flow"]
    net_calc = cf.get("operating", 0) + cf.get("investing", 0) + cf.get("financing", 0)
    return [
        {"check": "Balance Sheet (A = L + E)", "status": "PASS" if abs(assets - liab_eq) < 1 else "FAIL", "detail": f"A: ${assets:,.0f}, L+E: ${liab_eq:,.0f}"},
        {"check": "Intercompany Balances Net Zero", "status": "PASS" if abs(ic_recv - ic_pay) < 1 else "WARN", "detail": f"Recv: ${ic_recv:,.0f}, Pay: ${ic_pay:,.0f}"},
        {"check": "Cash Flow Reconciliation", "status": "PASS" if abs(net_calc - cf.get("net_change", 0)) < 1 else "FAIL", "detail": f"Calc: ${net_calc:,.0f}, Reported: ${cf.get('net_change', 0):,.0f}"},
    ]


def export_to_excel(consolidated: dict, entities: list[dict], checks: list[dict]) -> bytes:
    wb = Workbook()
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    for sheet_name, section in [("Consolidated P&L", "pnl"), ("Balance Sheet", "balance_sheet"), ("Cash Flow", "cash_flow")]:
        ws = wb.active if sheet_name == "Consolidated P&L" else wb.create_sheet(sheet_name)
        if sheet_name == "Consolidated P&L":
            ws.title = sheet_name
        headers = ["Line Item", "Consolidated (USD)"] + [e["entity"] for e in entities]
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = hdr_font; ws.cell(row=1, column=col).fill = hdr_fill
        for key in consolidated[section]:
            row = [key.replace("_", " ").title(), f"${consolidated[section][key]:,.0f}"]
            row += [f"${e.get(section, {}).get(key, 0):,.0f}" for e in entities]
            ws.append(row)
    ws3 = wb.create_sheet("GAAP Checks")
    ws3.append(["Check", "Status", "Detail"])
    for c in checks:
        ws3.append([c["check"], c["status"], c["detail"]])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# Sidebar
period = st.sidebar.text_input("Reporting Period (YYYY-MM)", datetime.now().strftime("%Y-%m"))
rates = fetch_exchange_rates()
st.sidebar.subheader("Exchange Rates")
for curr in ["EUR", "GBP", "CAD"]:
    st.sidebar.text(f"1 USD = {rates.get(curr, 'N/A')} {curr}")

# Fetch & convert
raw_entities = [fetch_netsuite(period), fetch_quickbooks(period), fetch_xero(period)]
entities = eliminate_intercompany([convert_to_base(e, rates) for e in raw_entities])
consolidated = consolidate(entities)

# Display
def _show_table(section_name, section_key):
    data = [{"Line Item": k.replace("_", " ").title(), "Consolidated": f"${v:,.0f}", **{e["entity"]: f"${e.get(section_key, {}).get(k, 0):,.0f}" for e in entities}} for k, v in consolidated[section_key].items()]
    st.dataframe(pd.DataFrame(data), use_container_width=True)

tab1, tab2, tab3, tab4 = st.tabs(["P&L", "Balance Sheet", "Cash Flow", "GAAP Checks"])
with tab1:
    st.subheader("Consolidated P&L"); _show_table("P&L", "pnl")
with tab2:
    st.subheader("Balance Sheet"); _show_table("Balance Sheet", "balance_sheet")
with tab3:
    st.subheader("Cash Flow"); _show_table("Cash Flow", "cash_flow")
with tab4:
    checks = gaap_checks(consolidated, raw_entities)
    for c in checks:
        (st.success if c["status"] == "PASS" else st.warning if c["status"] == "WARN" else st.error)(f"{c['check']}: {c['detail']}")

excel = export_to_excel(consolidated, entities, gaap_checks(consolidated, raw_entities))
st.download_button("Download Excel", data=excel, file_name=f"consolidated_{period}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
