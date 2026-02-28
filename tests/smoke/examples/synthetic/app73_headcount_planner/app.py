"""
Headcount Planner — Workforce planning and cost modeling.
Department heads use this to plan hires and compare against budget.
TODO: add role-based access so dept heads only see their own plans
"""

import streamlit as st
import requests
import json
from datetime import datetime, date
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill

# Workday API for current headcount
WORKDAY_API_URL = "https://wd5-services1.workday.com/ccx/api/v1/acme_corp"
WORKDAY_CLIENT_ID = "NjJhMzRkN2UtOGI5Yi00YzM2LWE4ZTktMGRiNmY3ZDgxNGUw"
WORKDAY_CLIENT_SECRET = "aG4rTm9SZW1vdGVfV29ya2RheV8yMDI0X3NlY3JldF9rZXlfcHJvZA=="
WORKDAY_REFRESH_TOKEN = "rt-workday-prod-8f7e6d5c4b3a2190-acme-hcm"

# Benefits multiplier — HR confirmed this is 22% loaded
BENEFITS_RATE = 0.22
EQUIPMENT_COST = 3000  # one-time per hire
ONSITE_MONTHLY = 500
REMOTE_MONTHLY = 100

# Comp bands by level (base salary) — from last comp cycle
# TODO: these are 2024 bands, need to update for 2025 cycle
COMP_BANDS = {
    "IC1": {"min": 70000, "mid": 85000, "max": 100000},
    "IC2": {"min": 90000, "mid": 110000, "max": 130000},
    "IC3": {"min": 120000, "mid": 145000, "max": 170000},
    "IC4": {"min": 150000, "mid": 180000, "max": 210000},
    "IC5": {"min": 190000, "mid": 225000, "max": 260000},
    "M1": {"min": 130000, "mid": 155000, "max": 180000},
    "M2": {"min": 165000, "mid": 195000, "max": 225000},
    "Dir": {"min": 200000, "mid": 240000, "max": 280000},
    "VP": {"min": 260000, "mid": 310000, "max": 360000},
}

# Approved budgets from Finance (FY25)
APPROVED_BUDGETS = {
    "Engineering": 4200000,
    "Product": 1800000,
    "Sales": 2500000,
    "Marketing": 1200000,
    "Customer Success": 950000,
    "G&A": 800000,
}

SCENARIO_MULTIPLIERS = {
    "Conservative": 0.75,
    "Base": 1.0,
    "Aggressive": 1.25,
}

def get_workday_headcount(department):
    """Fetch current headcount from Workday — simulated"""
    # TODO: implement actual Workday API call with OAuth2
    headers = {
        "Authorization": f"Bearer {WORKDAY_CLIENT_SECRET}",
        "Content-Type": "application/json",
    }
    simulated_headcount = {
        "Engineering": 45, "Product": 12, "Sales": 28,
        "Marketing": 15, "Customer Success": 10, "G&A": 18,
    }
    return simulated_headcount.get(department, 0)

def calculate_loaded_cost(base_salary, location_type, months_remaining):
    """Calculate fully loaded annual cost for a hire"""
    benefits = base_salary * BENEFITS_RATE
    monthly_office = ONSITE_MONTHLY if location_type == "On-site" else REMOTE_MONTHLY
    office_cost = monthly_office * months_remaining
    total = (base_salary + benefits) * (months_remaining / 12) + EQUIPMENT_COST + office_cost
    return round(total, 2)

def export_to_excel(plan_data, department, scenario):
    """Export headcount plan to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{department} - {scenario}"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    headers = ["Role", "Level", "Location", "Start Date", "Headcount", "Base Salary", "Loaded Cost"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, item in enumerate(plan_data, 2):
        ws.cell(row=row_idx, column=1, value=item["role"])
        ws.cell(row=row_idx, column=2, value=item["level"])
        ws.cell(row=row_idx, column=3, value=item["location"])
        ws.cell(row=row_idx, column=4, value=item["start_date"])
        ws.cell(row=row_idx, column=5, value=item["headcount"])
        ws.cell(row=row_idx, column=6, value=item["base_salary"])
        ws.cell(row=row_idx, column=7, value=item["loaded_cost"])

    total_row = len(plan_data) + 2
    ws.cell(row=total_row, column=6, value="TOTAL:").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=sum(i["loaded_cost"] for i in plan_data)).font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def main():
    st.set_page_config(page_title="Headcount Planner", layout="wide")
    st.title("Workforce Planning Tool")

    department = st.sidebar.selectbox("Department", list(APPROVED_BUDGETS.keys()))
    scenario = st.sidebar.selectbox("Scenario", list(SCENARIO_MULTIPLIERS.keys()))
    current_hc = get_workday_headcount(department)
    budget = APPROVED_BUDGETS[department]

    st.sidebar.metric("Current Headcount", current_hc)
    st.sidebar.metric("Approved Budget", f"${budget:,.0f}")

    st.subheader(f"Plan New Hires — {department} ({scenario})")

    if "plan_rows" not in st.session_state:
        st.session_state.plan_rows = []

    with st.form("add_role"):
        col1, col2, col3, col4 = st.columns(4)
        role = col1.text_input("Role Title", "Software Engineer")
        level = col2.selectbox("Level", list(COMP_BANDS.keys()))
        location = col3.selectbox("Location", ["On-site", "Remote", "Hybrid"])
        start_date = col4.date_input("Start Date", value=date(2025, 4, 1))
        headcount = st.number_input("Headcount", min_value=1, max_value=20, value=1)

        if st.form_submit_button("Add to Plan"):
            base = COMP_BANDS[level]["mid"]
            months_left = max(1, 12 - start_date.month + 1)
            loaded = calculate_loaded_cost(base, location, months_left) * headcount
            adjusted = loaded * SCENARIO_MULTIPLIERS[scenario]
            st.session_state.plan_rows.append({
                "role": role, "level": level, "location": location,
                "start_date": start_date.isoformat(), "headcount": headcount,
                "base_salary": base * headcount, "loaded_cost": round(adjusted, 2),
            })

    if st.session_state.plan_rows:
        st.dataframe(st.session_state.plan_rows)
        total_cost = sum(r["loaded_cost"] for r in st.session_state.plan_rows)
        total_hc = sum(r["headcount"] for r in st.session_state.plan_rows)

        col1, col2, col3 = st.columns(3)
        col1.metric("Total New Headcount", total_hc)
        col2.metric("Total Cost", f"${total_cost:,.0f}")
        col3.metric("Budget Remaining", f"${budget - total_cost:,.0f}",
                    delta=f"{'Over' if total_cost > budget else 'Under'} budget")

        if total_cost > budget:
            st.error(f"Plan exceeds approved budget by ${total_cost - budget:,.0f}")

        if st.button("Export to Excel"):
            buffer = export_to_excel(st.session_state.plan_rows, department, scenario)
            st.download_button("Download Plan", buffer, f"headcount_plan_{department}_{scenario}.xlsx",
                             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    main()
